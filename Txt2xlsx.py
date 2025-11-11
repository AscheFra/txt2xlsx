# -*- coding: utf-8 -*-
import os
import glob
import pandas as pd
import csv
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def try_read_lines(path):
    """Versuche, die Datei mit mehreren Encodings zu lesen und gebe die Zeilen zurück."""
    encodings = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
    for enc in encodings:
        try:
            with open(path, "r", encoding=enc, errors="replace") as f:
                return f.read().splitlines()
        except Exception:
            continue
    raise IOError("Datei kann nicht gelesen werden mit bekannten Encodings.")


def find_header_index(lines):
    """Finde den Index der Kopfzeile, in der eines der Felder genau 'Date' ist.

    Versucht mehrere mögliche Trennzeichen (Tab, Semikolon, Komma) und verwendet
    das csv-Modul, um korrekt mit Quotes und eingebetteten Kommas umzugehen.
    Gibt (index, header_list_from_date_onwards, delimiter, start_col) zurück oder (None, None, None, None).
    """
    delimiters = ["\t", ";", ","]
    for i, line in enumerate(lines):
        for d in delimiters:
            try:
                parts = next(csv.reader([line], delimiter=d))
            except Exception:
                parts = [p.strip() for p in line.split(d)]
            parts = [p.strip() for p in parts]
            # suche das Feld 'Date' irgendwo in parts
            if "Date" in parts:
                start_col = parts.index("Date")
                header = parts[start_col:]
                return i, header, d, start_col
    return None, None, None, None


def build_dataframe_from_lines(header, data_lines, delimiter="\t", start_col=0):
    """Erzeuge DataFrame aus Header (Liste) und data_lines (Liste von Zeilen).

    Nutzt das übergebene `delimiter` zum Splitten der Zeilen und `start_col`,
    falls die Date-Spalte nicht an Stelle 0 der gesplitteten Zeilen steht.
    Verwendet das csv-Modul, um korrekt mit Quotes und eingebetteten Delimitern umzugehen.
    """
    rows = []
    ncols = len(header)
    for ln in data_lines:
        try:
            parts = next(csv.reader([ln], delimiter=delimiter))
        except Exception:
            parts = [p.strip() for p in ln.split(delimiter)]
        parts = [p.strip() for p in parts]
        # slice starting at start_col
        sel = parts[start_col:start_col + ncols] if len(parts) > start_col else [""] * ncols
        # pad or trim
        if len(sel) < ncols:
            sel = sel + [""] * (ncols - len(sel))
        elif len(sel) > ncols:
            sel = sel[:ncols]
        rows.append(sel)
    df = pd.DataFrame(rows, columns=header)
    return df


def merge_date_time_if_present(df):
    """Wenn eine Time-Spalte existiert (z.B. 'Time(s)' oder 'Time (s)' o.ä.), entferne sie.

    Die Funktion sucht eine Spalte, deren normalisierter Name mit 'time' beginnt
    und entfernt diese Spalte vollständig aus dem DataFrame.
    """
    def find_time_col(columns):
        for c in columns:
            if not isinstance(c, str):
                continue
            norm = ''.join(ch.lower() for ch in c if ch.isalnum())
            if norm.startswith('time'):
                return c
        return None

    if 'Date' not in df.columns:
        return df

    time_col = find_time_col(df.columns)
    if time_col and time_col != 'Date':
        if time_col in df.columns:
            df = df.drop(columns=[time_col])
            print(f"Entferne erkannte Time-Spalte: '{time_col}'")
    return df


def _normalize_number_str(s) -> str:
    """Normalisiere eine numerische Zeichenkette zu einem string (niemals None):
    - Entferne Tausendertrennzeichen '.' falls vorhanden (z.B. '1.234' -> '1234')
    - Ersetze Dezimalkomma mit Punkt (',' -> '.')
    - Trimme whitespace
    """
    if s is None:
        return ''
    t = str(s).strip()
    # Wenn sowohl '.' als Tausender und ',' als Dezimal vorkommen, entferne Tausenderpunkte
    if t.count('.') > 1 and ',' in t:
        t = t.replace('.', '')
    # Standard: ersetze Komma durch Punkt
    t = t.replace(',', '.')
    return t


def _to_numeric_series(ser):
    """Versuche, eine Serie von Strings in numerische Werte zu konvertieren nachdem normalisiert wurde.

    Verwende eine List-Comprehension und baue daraus eine pandas.Series, das vermeidet statische Analyzer-Warnungen
    wenn der Typ von `ser` nicht exakt erkannt wird.
    """
    try:
        arr = [ _normalize_number_str(x) for x in ser.astype(str) ]
    except Exception:
        # Fallback: versuche, iterierbar direkt zu verarbeiten
        arr = [ _normalize_number_str(x) for x in ser ]
    s = pd.Series(arr)
    return pd.to_numeric(s, errors='coerce')


def find_start_index_by_penult_col(df):
    """Kompatibilitäts-Wrapper: wie vorher, verwende die vorletzte Spalte (offset=2).

    Intern wird nun `find_start_index_by_offset` verwendet.
    """
    return find_start_index_by_offset(df, offset_from_right=2)


def find_start_index_by_offset(df, offset_from_right=2):
    """Bestimme den ersten Index, bei dem die Spalte `offset_from_right` von rechts != 0 (nach Numerisierung).

    offset_from_right: 1 = letzte Spalte, 2 = vorletzte (default früher), 4 = viertletzte (für .his).
    Liefert (idx, column_name) oder (None, column_name).
    """
    if df.shape[1] < offset_from_right:
        return None, None
    col_idx = -offset_from_right
    try:
        penult_col = df.columns[col_idx]
    except Exception:
        return None, None
    ser_num = _to_numeric_series(df[penult_col])
    mask = ser_num.notna() & (ser_num != 0)
    if mask.any():
        idx = int(mask.idxmax())
        return idx, penult_col
    # Fallback: scanne andere Spalten (außer 'Date'), von rechts nach links
    for col in df.columns[::-1]:
        if col == 'Date':
            continue
        ser_num = _to_numeric_series(df[col])
        mask = ser_num.notna() & (ser_num != 0)
        if mask.any():
            idx = int(mask.idxmax())
            return idx, col
    return None, penult_col


def find_last_index_by_offset(df, offset_from_right=2):
    """Bestimme den letzten Index (von oben gezählt), bei dem die Spalte `offset_from_right` von rechts != 0 ist.

    Liefert (idx, column_name) oder (None, column_name).
    """
    if df.shape[1] < offset_from_right:
        return None, None
    try:
        target_col = df.columns[-offset_from_right]
    except Exception:
        return None, None
    ser_num = _to_numeric_series(df[target_col])
    mask = ser_num.notna() & (ser_num != 0)
    if mask.any():
        # letzter True-Index finden
        true_indices = list(mask[mask].index)
        last_idx = int(true_indices[-1])
        return last_idx, target_col
    # Fallback: scanne andere Spalten (außer 'Date'), von rechts nach links
    for col in df.columns[::-1]:
        if col == 'Date':
            continue
        ser_num = _to_numeric_series(df[col])
        mask = ser_num.notna() & (ser_num != 0)
        if mask.any():
            true_indices = list(mask[mask].index)
            last_idx = int(true_indices[-1])
            return last_idx, col
    return None, target_col


def find_trim_indices_by_offset(df, offset_from_right=2):
    """Gebe (start_idx, end_idx, used_column) zurück.

    start_idx: erster Index von oben mit Wert != 0 in der relevanten Spalte
    end_idx: letzter Index (von oben gezählt) mit Wert != 0 in der relevanten Spalte
    used_column: Name der verwendeten Spalte
    Wenn keine passende Spalte/kein Wert gefunden wird, liefert (None, None, column_name).
    """
    if df.shape[1] < offset_from_right:
        return None, None, None
    try:
        col = df.columns[-offset_from_right]
    except Exception:
        return None, None, None
    ser_num = _to_numeric_series(df[col])
    mask = ser_num.notna() & (ser_num != 0)
    if mask.any():
        # idxmin/idxmax nicht zuverlässig bei nicht-boolean Indexen -> nutze list
        true_idxs = list(mask[mask].index)
        start_idx = int(true_idxs[0])
        end_idx = int(true_idxs[-1])
        return start_idx, end_idx, col
    # Fallback: scanne andere Spalten (außer 'Date'), von rechts nach links
    for c in df.columns[::-1]:
        if c == 'Date':
            continue
        ser_num = _to_numeric_series(df[c])
        mask = ser_num.notna() & (ser_num != 0)
        if mask.any():
            true_idxs = list(mask[mask].index)
            start_idx = int(true_idxs[0])
            end_idx = int(true_idxs[-1])
            return start_idx, end_idx, c
    return None, None, col


def get_input_path_from_user():
    """Öffnet zwingend einen Datei-Auswahl-Dialog für .txt/.csv/.his.

    Falls tkinter nicht verfügbar ist (z. B. Headless-Umgebung), fällt die Funktion
    auf die ursprüngliche Logik zurück und nimmt die erste gefundene .txt/.csv/.his
    im aktuellen Ordner.
    """
    # Versuche, einen grafischen Datei-Dialog zu öffnen
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        filetypes = [("Text/CSV/HIS", ("*.txt", "*.csv", "*.his", "*.HIS")), ("All files", "*.*")]
        path = filedialog.askopenfilename(title="Wähle .txt, .csv oder .his Datei", filetypes=filetypes)
        root.destroy()
        if not path:
            print("Keine Datei ausgewählt.")
            return None
        print(f"Gewählte Datei: {path}")
        return path
    except Exception as e:
        # Fallback: erste gefundene Datei wie vorher
        print("GUI-Dateidialog nicht verfügbar, verwende erstes gefundenes File (Fallback).", e)
        txts = sorted(glob.glob("*.txt"))
        csvs = sorted(glob.glob("*.csv"))
        his1 = sorted(glob.glob("*.his"))
        his2 = sorted(glob.glob("*.HIS"))
        his = his1 + his2
        candidates = txts + csvs + his
        if not candidates:
            print("Keine .txt-, .csv- oder .his-Datei im aktuellen Verzeichnis gefunden.")
            return None
        print(f"Verwende erste gefundene Datei: {candidates[0]}")
        return candidates[0]


def get_sampling_k():
    """Frage per einfacherem GUI-Dialog nach dem Sampling-K (Integer >=1).

    Falls tkinter nicht verfügbar ist, wird der Default 1 ohne Terminalabfrage zurückgegeben.
    """
    try:
        import tkinter as tk
        from tkinter import simpledialog
        root = tk.Tk()
        root.withdraw()
        k = simpledialog.askinteger("Sampling", "Jeden wievielten Datenpunkt übernehmen?", initialvalue=1, minvalue=1)
        root.destroy()
        if k is None:
            print("Keine Eingabe im Dialog, verwende Default 1.")
            return 1
        return int(k)
    except Exception:
        # Kein GUI verfügbar -> verwende Default 1 (keine Terminalabfrage)
        print("GUI für Sampling-Dialog nicht verfügbar, verwende Default 1.")
        return 1


def convert_df_numbers_to_german_strings(df, date_cols=('Date',)):
    """Konvertiert numerische Zellen in Strings mit Komma als Dezimaltrennzeichen.

    - date_cols: Spaltennamen, die nicht umformatiert werden sollen (z.B. 'Date').
    - Nicht-numerische Zellen bleiben unverändert.
    - Rückgabe: neuer DataFrame (Kopien der Spalten, numerische Werte als Strings).
    """
    df2 = df.copy()
    # einfache Regex für eine reine numerische Schreibweise (mit Punkt oder Komma)
    num_re = re.compile(r'^[+-]?\d+[.,]?\d*$')
    for col in df2.columns:
        if col in date_cols:
            continue
        col_vals = df2[col].astype(str)
        new_vals = []
        for orig in col_vals:
            s = orig.strip()
            if s == '' or s.lower() in ('nan', 'none'):
                new_vals.append('')
                continue
            # Wenn der Originalstring eine einfache Zahl darstellt (z.B. -15.83 oder 0.00)
            if num_re.match(s):
                # Tausenderpunkte entfernen falls vorhanden (z.B. 1.234 -> 1234)
                # aber nur wenn sowohl '.' und ',' vorkommen oder mehrere Punkte
                if s.count('.') > 1 and ',' in s:
                    s2 = s.replace('.', '')
                else:
                    s2 = s
                # Ersetze Dezimalpunkt durch Komma
                s2 = s2.replace('.', ',')
                new_vals.append(s2)
                continue
            # Falls nicht-matching, versuche numerisch zu parsen und formatiere dann
            try:
                f = float(s.replace(',', '.'))
            except Exception:
                new_vals.append(orig)
                continue
            # Bestimme Dezimalstellen basierend auf Originalstring, falls vorhanden
            dec = None
            if '.' in s:
                dec = len(s.split('.')[-1])
            elif ',' in s:
                dec = len(s.split(',')[-1])
            if dec is not None:
                fmt = '{:.' + str(dec) + 'f}'
                formatted = fmt.format(f).replace('.', ',')
            else:
                # Keine Information, schreibe kompaktes Format
                if f.is_integer():
                    formatted = str(int(f))
                else:
                    # verwende repr, aber ersetze '.'->','
                    formatted = str(f).replace('.', ',')
            new_vals.append(formatted)
        df2[col] = new_vals
    return df2


def write_df_to_excel_with_formats(path, df, date_cols=('Date',), max_decimals_cap=6):
    """Schreibe DataFrame nach Excel und setze Number-Format für numerische Spalten.

    - Konvertiert alle nicht-`date_cols`-Spalten in numerische Werte (NaN falls nicht konvertierbar).
    - Schreibt mit pandas.to_excel und öffnet die Datei anschließend mit openpyxl, um die Number-Format-Codes
      für numerische Zellen zu setzen (Anzahl Dezimalstellen wird aus den Daten bestimmt, begrenzt durch max_decimals_cap).
    - Excel zeigt Dezimaltrennzeichen entsprechend der Benutzer-Regional-Einstellungen (deutsche Excel-Instanz zeigt Komma).
    """
    df_copy = df.copy()
    # Konvertiere Spalten in numerisch, wo möglich
    for col in df_copy.columns:
        if col in date_cols:
            continue
        try:
            df_copy[col] = _to_numeric_series(df_copy[col])
        except Exception:
            # falls konvertierung fehlschlägt, belasse Spalte unverändert
            continue

    # Schreibe zuerst mit pandas (openpyxl engine)
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        df_copy.to_excel(writer, index=False)

    # Öffne mit openpyxl und setze number_format für numerische Spalten
    wb = load_workbook(path)
    ws = wb.active
    nrows = ws.max_row

    for idx, col in enumerate(df_copy.columns, start=1):
        header = col
        if header in date_cols:
            continue
        ser = df_copy[header]
        # bestimme, ob Spalte numerisch (mindestens ein numerischer Wert)
        nums = ser.dropna().astype(float)
        if nums.empty:
            continue
        # bestimme maximale Anzahl Dezimalstellen in Daten (bis max_decimals_cap)
        max_dec = 0
        for v in nums:
            if pd.isna(v):
                continue
            # arbeite mit String-Repräsentation mit begrenzter Genauigkeit
            s = ('{:.%df}' % max_decimals_cap).format(v) if False else ('{:.6f}'.format(v))
            s = s.rstrip('0').rstrip('.')
            if '.' in s:
                dec = len(s.split('.')[-1])
                if dec > max_dec:
                    max_dec = dec
            if max_dec >= max_decimals_cap:
                break
        # wähle Format
        if max_dec == 0:
            fmt = '0'
        else:
            fmt = '0.' + ('0' * max_dec)
        col_letter = get_column_letter(idx)
        # setze number_format für alle numerischen Zellen in der Spalte
        for row in range(2, nrows + 1):
            cell = ws[f"{col_letter}{row}"]
            if isinstance(cell.value, (int, float)):
                cell.number_format = fmt
    wb.save(path)


def main():
    print("Starte Konvertierung txt/csv/his -> xlsx")
    path = get_input_path_from_user()
    if not path:
        return
    if not os.path.isfile(path):
        print(f"Datei nicht gefunden: {path}")
        return

    k = get_sampling_k()

    try:
        lines = try_read_lines(path)
    except Exception as e:
        print("Fehler beim Lesen der Datei:", e)
        return

    header_idx, header, delimiter, start_col = find_header_index(lines)
    if header_idx is None:
        print("Kopfzeile mit 'Date' nicht gefunden. Abbruch.")
        return

    data_lines = lines[header_idx + 1 :]
    if not data_lines:
        print("Keine Datenzeilen nach der Kopfzeile gefunden. Abbruch.")
        return

    df = build_dataframe_from_lines(header, data_lines, delimiter, start_col)

    # Time(s) behandeln: an Date anhängen und entfernen
    df = merge_date_time_if_present(df)

    if df.empty:
        print("Keine Daten vorhanden nach Verarbeitung.")
        return

    # Bestimme Trim-Indices (Start/Ende) basierend auf Dateityp
    if path.lower().endswith('.his'):
        offset = 4
    else:
        offset = 2

    start_idx, end_idx, used_col = find_trim_indices_by_offset(df, offset_from_right=offset)
    if start_idx is None or end_idx is None:
        print(f"Kein Wert != 0 in der relevanten Spalte (offset={offset}) gefunden. Abbruch.")
        return

    # Trunkiere die Tabelle zwischen start_idx und end_idx (inklusive)
    truncated = df.iloc[start_idx : end_idx + 1]
    # Sampling: jedes k-te Element aus der getrimmten Tabelle
    out_df = truncated.iloc[::k].reset_index(drop=True)

    # Ausgabe-Dateiname: gleicher Basisname mit .xlsx
    base = os.path.splitext(os.path.basename(path))[0]
    out_name = base + ".xlsx"
    if os.path.exists(out_name):
        print(f"Warnung: Zieldatei '{out_name}' existiert bereits. Datei wird nicht überschrieben.")
        return

    try:
        write_df_to_excel_with_formats(out_name, out_df, date_cols=('Date',))
        print(f"Erfolgreich gespeichert: {out_name}")
        print(f"Zeilenbereich (0-basiert im DataFrame): {start_idx} .. {end_idx}, Sampling: jede {k}. Zeile (entscheidende Spalte: {used_col})")
    except Exception as e:
        print("Fehler beim Speichern der Excel-Datei:", e)


if __name__ == "__main__":
    main()
